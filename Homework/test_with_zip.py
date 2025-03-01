import csv
from zipfile import ZipFile

from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader

from Homework.conftest import ARCHIVE_FILE_PATH


def test_pdf():
    print(f"Содержимое архива {ARCHIVE_FILE_PATH}:")
    with ZipFile(ARCHIVE_FILE_PATH) as zip_file:
        for name in zip_file.namelist():
            print(f"  {name}")

    with ZipFile(ARCHIVE_FILE_PATH) as zip_file:
        with zip_file.open("Python Testing with Pytest (Brian Okken).pdf") as pdf:
            pdf_reader = PdfReader(pdf)
            count_of_pages = len(pdf_reader.pages)

            assert count_of_pages > 0
            assert 'Pragmatic Bookshelf' in pdf_reader.pages[2].extract_text()


def test_xlsx():
    with ZipFile(ARCHIVE_FILE_PATH) as zip_file:
        with zip_file.open('file_example_XLSX_50.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            assert sheet.cell(row=50, column=7).value == '15/10/2017'


def test_csv():
    with ZipFile(ARCHIVE_FILE_PATH) as zip_file:
        with zip_file.open("Tin-fuid-13a0f23b8b45f14e-guest-full.csv") as csv_file:
            content = csv_file.read().decode('utf-8')
            csv_reader = list(csv.reader(content.splitlines(), delimiter=','))
            first_row_values = csv_reader[0]
            assert first_row_values[0] == '# event;subevent;status;phid;time;geo;pos dist; time dist;speed;user0'
            assert len(csv_reader) == 1091
